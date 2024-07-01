from openpyxl import Workbook,load_workbook

src_filename = 'in.xlsx'
dst_filename = 'out.xlsx'

def main():
    # in
    wbi = load_workbook(filename=src_filename)
    wsi = wbi.active
    maxRows = wsi.max_row
    maxColumns = wsi.max_column
    # out
    wbo = Workbook()
    wso = wbo.active
    wso.title = wsi.title
    # title
    cells_first = [wsi.cell(1, j).value for j in range(1, maxColumns + 1)]
    no_index = cells_first.index('No.')
    des_index = cells_first.index('Designator')
    qua_index = cells_first.index('Quantity')
    wso.append(cells_first)

    delta = 0
    for i in range(2,maxRows+1):
        value = [wsi.cell(i, j).value for j in range(1, maxColumns + 1)]
        if value[des_index] == None:
            break
        # Len of char is limitted to less than 200
        if len(value[des_index]) >= 200:
            quantity = 1
            unitstr = ''
            # split
            for v in value[des_index].split(','):
                if len(unitstr) + len(v) > 199:
                    value[no_index] = str(delta + i - 1)
                    value[qua_index] = str(quantity)
                    value[des_index] = unitstr
                    wso.append(value)
                    unitstr = v
                    quantity = 1
                    delta += 1
                else:
                    if unitstr == '':
                        unitstr = v
                    else:
                        unitstr = unitstr + ',' + v
                    quantity += 1
            
            value[no_index] = str(delta + i - 1)
            value[qua_index] = str(quantity)
            value[des_index] = unitstr
            wso.append(value)
            unitstr = v
        else:
            value[no_index] = str(int(value[0]) + delta)
            wso.append(value)

    wbo.save(filename = dst_filename)

if __name__ == '__main__':
    main()